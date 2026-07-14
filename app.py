import os
import io
import json
import traceback
from datetime import datetime, timedelta
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, Response
)
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash
from database import init_db, query_db, execute_db, get_next_order_code, encode_image

app = Flask(__name__)
app.secret_key = 'hokiapp-secret-key-2024-change-in-production'
app.config['MAX_CONTENT_LENGTH'] = 8 * 1024 * 1024  # 8MB max upload

ALLOWED_IMAGE_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


def allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXT


# ─── Auth Decorators ────────────────────────────────────────────────────────

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return jsonify({'error': 'Akses ditolak. Hanya admin yang dapat melakukan aksi ini.'}), 403
        return f(*args, **kwargs)
    return decorated


# ─── Context Processor (inject settings into all templates) ─────────────────

@app.context_processor
def inject_globals():
    """Inject global settings and user info into every template context."""
    settings = {}
    try:
        rows = query_db("SELECT key, value FROM settings")
        settings = {r['key']: r['value'] for r in rows}
    except Exception:
        pass
    return {
        'app_settings': settings,
        'current_user': {
            'id':           session.get('user_id'),
            'username':     session.get('username'),
            'role':         session.get('role'),
            'nama_lengkap': session.get('nama_lengkap'),
        }
    }


@app.errorhandler(Exception)
def handle_exception(e):
    """Global error handler to write exceptions to logsystem.txt."""
    if isinstance(e, HTTPException):
        return e
        
    try:
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logsystem.txt')
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"\n{'='*50}\n")
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ERROR in {request.method} {request.url}\n")
            f.write(traceback.format_exc())
    except:
        pass
        
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Terjadi kesalahan internal. Cek logsystem.txt'}), 500
    return "Terjadi Kesalahan Internal. Silakan cek logsystem.txt pada folder aplikasi.", 500


# ─── AUTH ROUTES ────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = query_db(
            "SELECT * FROM users WHERE username = ? AND is_active = 1",
            [username], one=True
        )
        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id']     = user['id']
            session['username']    = user['username']
            session['role']        = user['role']
            session['nama_lengkap']= user['nama_lengkap']
            session.permanent      = True
            return redirect(url_for('dashboard'))
        else:
            error = 'Username atau password salah.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ─── PAGE ROUTES ────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


@app.route('/kasir')
@login_required
def kasir():
    return render_template('kasir.html')


@app.route('/payment')
@login_required
def payment():
    return render_template('payment.html')


@app.route('/stock')
@login_required
def stock():
    return render_template('stock.html')


@app.route('/laporan')
@login_required
def laporan():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    return render_template('laporan.html')


@app.route('/settings')
@login_required
def settings():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    return render_template('settings.html')


# ─── API: CATEGORIES ────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['GET'])
@login_required
def api_categories():
    cats = query_db("SELECT * FROM categories ORDER BY sort_order, name")
    return jsonify([dict(c) for c in cats])


@app.route('/api/categories', methods=['POST'])
@admin_required
def api_categories_create():
    data = request.json or {}
    name = data.get('name', '').strip()
    icon = data.get('icon', 'category')
    desc = data.get('description', '')
    sort = int(data.get('sort_order', 0))
    if not name:
        return jsonify({'error': 'Nama kategori wajib diisi.'}), 400
    cat_id = execute_db(
        "INSERT INTO categories (name, icon, description, sort_order) VALUES (?, ?, ?, ?)",
        [name, icon, desc, sort]
    )
    return jsonify({'id': cat_id, 'name': name, 'icon': icon, 'description': desc}), 201


@app.route('/api/categories/<int:cat_id>', methods=['PUT'])
@admin_required
def api_categories_update(cat_id):
    data = request.json or {}
    existing = query_db("SELECT * FROM categories WHERE id = ?", [cat_id], one=True)
    if not existing:
        return jsonify({'error': 'Kategori tidak ditemukan.'}), 404
    execute_db(
        "UPDATE categories SET name=?, icon=?, description=?, sort_order=? WHERE id=?",
        [data.get('name', existing['name']),
         data.get('icon', existing['icon']),
         data.get('description', existing['description']),
         int(data.get('sort_order', existing['sort_order'])),
         cat_id]
    )
    return jsonify({'message': 'Kategori diperbarui.'})


@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
@admin_required
def api_categories_delete(cat_id):
    execute_db("DELETE FROM categories WHERE id = ?", [cat_id])
    return jsonify({'message': 'Kategori dihapus.'})


# ─── API: PRODUCTS ──────────────────────────────────────────────────────────

@app.route('/api/products', methods=['GET'])
@login_required
def api_products():
    category_id = request.args.get('category_id')
    search      = request.args.get('search', '').strip()
    featured    = request.args.get('featured')

    query = """
        SELECT p.id, p.name, p.description, p.price, p.hpp, p.discount_pct,
               p.stock_qty, p.stock_unit, p.low_stock_threshold,
               p.is_available, p.is_featured, p.created_at,
               CASE WHEN p.image_data IS NOT NULL AND p.image_data != '' THEN 1 ELSE 0 END as has_image,
               c.name as category_name, c.icon as category_icon, p.category_id
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE 1=1
    """
    args = []
    if category_id and category_id != 'all':
        query += " AND p.category_id = ?"
        args.append(category_id)
    if search:
        query += " AND p.name LIKE ?"
        args.append(f'%{search}%')
    if featured == '1':
        query += " AND p.is_featured = 1"
    query += " ORDER BY c.sort_order, p.name"

    products = query_db(query, args)
    return jsonify([dict(p) for p in products])


@app.route('/api/products/<int:prod_id>/image', methods=['GET'])
@login_required
def api_product_image(prod_id):
    """Serve product image stored as base64 in DB."""
    row = query_db("SELECT image_data FROM products WHERE id = ?", [prod_id], one=True)
    if not row or not row['image_data']:
        return '', 404
    data_url = row['image_data']
    # data_url format: "data:image/jpeg;base64,<b64>"
    try:
        header, b64 = data_url.split(',', 1)
        mime = header.split(':')[1].split(';')[0]
        import base64
        raw = base64.b64decode(b64)
        return Response(raw, mimetype=mime)
    except Exception:
        return '', 404


@app.route('/api/products', methods=['POST'])
@login_required
def api_products_create():
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Nama produk wajib diisi.'}), 400

    try:
        price     = float(request.form.get('price', 0))
        hpp       = float(request.form.get('hpp', 0))
        discount  = float(request.form.get('discount_pct', 0))
        stock_qty = int(request.form.get('stock_qty', 0))
        threshold = int(request.form.get('low_stock_threshold', 5))
    except (ValueError, TypeError):
        return jsonify({'error': 'Nilai numerik tidak valid.'}), 400

    category_id  = request.form.get('category_id') or None
    stock_unit   = request.form.get('stock_unit', 'porsi')
    description  = request.form.get('description', '')
    is_available = int(request.form.get('is_available', 1))
    is_featured  = int(request.form.get('is_featured', 0))
    image_data   = None

    if 'image' in request.files:
        file = request.files['image']
        if file and file.filename and allowed_image(file.filename):
            image_data = encode_image(file)

    prod_id = execute_db(
        """INSERT INTO products
           (category_id, name, description, price, hpp, discount_pct, stock_qty, stock_unit,
            low_stock_threshold, image_data, is_available, is_featured)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        [category_id, name, description, price, hpp, discount, stock_qty, stock_unit,
         threshold, image_data, is_available, is_featured]
    )

    # Log initial stock
    if stock_qty > 0:
        execute_db(
            """INSERT INTO stock_history (product_id, product_name, action, qty_change, qty_before, qty_after, user_id)
               VALUES (?, ?, 'initial', ?, 0, ?, ?)""",
            [prod_id, name, stock_qty, stock_qty, session.get('user_id')]
        )

    product = query_db(
        """SELECT p.id, p.name, p.description, p.price, p.hpp, p.discount_pct,
                  p.stock_qty, p.stock_unit, p.low_stock_threshold,
                  p.is_available, p.is_featured, p.category_id,
                  CASE WHEN p.image_data IS NOT NULL AND p.image_data != '' THEN 1 ELSE 0 END as has_image,
                  c.name as category_name
           FROM products p LEFT JOIN categories c ON p.category_id = c.id
           WHERE p.id = ?""",
        [prod_id], one=True
    )
    return jsonify(dict(product)), 201


@app.route('/api/products/<int:prod_id>', methods=['PUT'])
@login_required
def api_products_update(prod_id):
    existing = query_db("SELECT * FROM products WHERE id = ?", [prod_id], one=True)
    if not existing:
        return jsonify({'error': 'Produk tidak ditemukan.'}), 404

    image_data = existing['image_data']

    if request.content_type and 'multipart/form-data' in request.content_type:
        data = request.form
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename and allowed_image(file.filename):
                image_data = encode_image(file)
        if data.get('remove_image') == '1':
            image_data = None
    else:
        data = request.json or {}

    name = data.get('name', existing['name']).strip() if data.get('name') else existing['name']
    try:
        price     = float(data.get('price',           existing['price']))
        hpp       = float(data.get('hpp',              existing['hpp']))
        discount  = float(data.get('discount_pct',     existing['discount_pct']))
        stock_qty = int(data.get('stock_qty',          existing['stock_qty']))
        threshold = int(data.get('low_stock_threshold',existing['low_stock_threshold']))
    except (ValueError, TypeError):
        return jsonify({'error': 'Nilai tidak valid.'}), 400

    category_id  = data.get('category_id', existing['category_id']) or None
    stock_unit   = data.get('stock_unit',   existing['stock_unit'])
    description  = data.get('description',  existing['description'] or '')
    is_available = int(data.get('is_available', existing['is_available']))
    is_featured  = int(data.get('is_featured',  existing['is_featured']))

    execute_db(
        """UPDATE products SET
           category_id=?, name=?, description=?, price=?, hpp=?, discount_pct=?,
           stock_qty=?, stock_unit=?, low_stock_threshold=?, image_data=?,
           is_available=?, is_featured=?, updated_at=CURRENT_TIMESTAMP
           WHERE id=?""",
        [category_id, name, description, price, hpp, discount, stock_qty, stock_unit,
         threshold, image_data, is_available, is_featured, prod_id]
    )
    product = query_db(
        """SELECT p.id, p.name, p.description, p.price, p.hpp, p.discount_pct,
                  p.stock_qty, p.stock_unit, p.low_stock_threshold,
                  p.is_available, p.is_featured, p.category_id,
                  CASE WHEN p.image_data IS NOT NULL AND p.image_data != '' THEN 1 ELSE 0 END as has_image,
                  c.name as category_name
           FROM products p LEFT JOIN categories c ON p.category_id = c.id
           WHERE p.id = ?""",
        [prod_id], one=True
    )
    return jsonify(dict(product))


@app.route('/api/products/<int:prod_id>', methods=['DELETE'])
@admin_required
def api_products_delete(prod_id):
    existing = query_db("SELECT * FROM products WHERE id = ?", [prod_id], one=True)
    if not existing:
        return jsonify({'error': 'Produk tidak ditemukan.'}), 404
    execute_db("DELETE FROM products WHERE id = ?", [prod_id])
    return jsonify({'message': 'Produk berhasil dihapus.'})


@app.route('/api/products/<int:prod_id>/restock', methods=['POST'])
@login_required
def api_products_restock(prod_id):
    data = request.json or {}
    try:
        qty = int(data.get('qty', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Jumlah tidak valid.'}), 400
    if qty <= 0:
        return jsonify({'error': 'Jumlah harus lebih dari 0.'}), 400

    product = query_db("SELECT * FROM products WHERE id = ?", [prod_id], one=True)
    if not product:
        return jsonify({'error': 'Produk tidak ditemukan.'}), 404

    old_qty = product['stock_qty']
    new_qty = old_qty + qty
    execute_db("UPDATE products SET stock_qty = ?, updated_at=CURRENT_TIMESTAMP WHERE id = ?", [new_qty, prod_id])

    # Log stock history
    execute_db(
        """INSERT INTO stock_history (product_id, product_name, action, qty_change, qty_before, qty_after, notes, user_id)
           VALUES (?, ?, 'restock', ?, ?, ?, ?, ?)""",
        [prod_id, product['name'], qty, old_qty, new_qty, data.get('notes', ''), session.get('user_id')]
    )
    return jsonify({'message': 'Restock berhasil.', 'new_qty': new_qty})


@app.route('/api/products/<int:prod_id>/adjust', methods=['POST'])
@admin_required
def api_products_adjust(prod_id):
    """Manual stock adjustment (positive or negative)."""
    data = request.json or {}
    try:
        qty = int(data.get('qty', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Jumlah tidak valid.'}), 400

    product = query_db("SELECT * FROM products WHERE id = ?", [prod_id], one=True)
    if not product:
        return jsonify({'error': 'Produk tidak ditemukan.'}), 404

    old_qty = product['stock_qty']
    new_qty = max(0, old_qty + qty)
    execute_db("UPDATE products SET stock_qty = ?, updated_at=CURRENT_TIMESTAMP WHERE id = ?", [new_qty, prod_id])
    execute_db(
        """INSERT INTO stock_history (product_id, product_name, action, qty_change, qty_before, qty_after, notes, user_id)
           VALUES (?, ?, 'adjustment', ?, ?, ?, ?, ?)""",
        [prod_id, product['name'], qty, old_qty, new_qty, data.get('notes', 'Penyesuaian manual'), session.get('user_id')]
    )
    return jsonify({'message': 'Stok disesuaikan.', 'new_qty': new_qty})


# ─── API: TRANSACTIONS ──────────────────────────────────────────────────────

@app.route('/api/transactions', methods=['GET'])
@login_required
def api_transactions():
    period   = request.args.get('period', 'today')
    status   = request.args.get('status', 'all')
    limit    = int(request.args.get('limit', 50))
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')

    now = datetime.now()
    if date_from and date_to:
        from_str = date_from + ' 00:00:00'
        to_str   = date_to   + ' 23:59:59'
    elif period == 'today':
        from_str = now.strftime('%Y-%m-%d 00:00:00')
        to_str   = now.strftime('%Y-%m-%d 23:59:59')
    elif period == 'week':
        from_str = (now - timedelta(days=7)).strftime('%Y-%m-%d 00:00:00')
        to_str   = now.strftime('%Y-%m-%d 23:59:59')
    elif period == 'month':
        from_str = (now - timedelta(days=30)).strftime('%Y-%m-%d 00:00:00')
        to_str   = now.strftime('%Y-%m-%d 23:59:59')
    else:
        from_str = '2000-01-01 00:00:00'
        to_str   = '2099-12-31 23:59:59'

    query = """
        SELECT t.*,
               u.nama_lengkap as cashier_display_name,
               GROUP_CONCAT(ti.product_name || ' x' || ti.quantity, ', ') as items_summary,
               COUNT(ti.id) as items_count
        FROM transactions t
        LEFT JOIN users u ON t.cashier_id = u.id
        LEFT JOIN transaction_items ti ON t.id = ti.transaction_id
        WHERE t.created_at BETWEEN ? AND ?
    """
    args = [from_str, to_str]

    if status != 'all':
        query += " AND t.status = ?"
        args.append(status)

    query += " GROUP BY t.id ORDER BY t.created_at DESC LIMIT ?"
    args.append(limit)

    txns = query_db(query, args)
    return jsonify([dict(t) for t in txns])


@app.route('/api/transactions/pending', methods=['GET'])
@login_required
def api_transactions_pending():
    txns = query_db(
        """SELECT t.*, u.nama_lengkap as cashier_display_name
           FROM transactions t
           LEFT JOIN users u ON t.cashier_id = u.id
           WHERE t.status = 'pending'
           ORDER BY t.created_at DESC"""
    )
    return jsonify([dict(t) for t in txns])


@app.route('/api/transactions', methods=['POST'])
@login_required
def api_transactions_create():
    data  = request.json or {}
    items = data.get('items', [])
    if not items:
        return jsonify({'error': 'Pesanan tidak boleh kosong.'}), 400

    # Tax
    s = query_db("SELECT key, value FROM settings WHERE key IN ('tax_enabled','tax_rate')")
    sv = {r['key']: r['value'] for r in s}
    tax_enabled = sv.get('tax_enabled', '1') == '1'
    tax_rate    = float(sv.get('tax_rate', '10')) / 100

    subtotal = sum(item.get('price', 0) * item.get('qty', 1) for item in items)
    discount = float(data.get('discount_amount', 0))
    taxable  = max(0, subtotal - discount)
    tax_amount = round(taxable * tax_rate, 0) if tax_enabled else 0
    total    = taxable + tax_amount

    cashier = query_db("SELECT nama_lengkap FROM users WHERE id = ?", [session.get('user_id')], one=True)
    cashier_name = cashier['nama_lengkap'] if cashier else session.get('username')

    txn_id = execute_db(
        """INSERT INTO transactions
           (order_code, cashier_id, cashier_name, customer_name, table_number,
            subtotal, discount_amount, tax_amount, total, status, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)""",
        [get_next_order_code(), session.get('user_id'), cashier_name,
         data.get('customer_name', ''), data.get('table_number', ''),
         subtotal, discount, tax_amount, total, data.get('notes', '')]
    )

    for item in items:
        execute_db(
            """INSERT INTO transaction_items
               (transaction_id, product_id, product_name, product_price, quantity, discount_pct, subtotal, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            [txn_id, item.get('product_id'), item.get('name') or 'Produk', item.get('price', 0),
             item.get('qty', 1), item.get('discount_pct', 0),
             item.get('price') * item.get('qty', 1), item.get('notes', '')]
        )

    order = query_db("SELECT order_code FROM transactions WHERE id = ?", [txn_id], one=True)
    return jsonify({'id': txn_id, 'order_code': order['order_code'], 'total': total}), 201


@app.route('/api/transactions/<int:txn_id>', methods=['GET'])
@login_required
def api_transaction_detail(txn_id):
    txn = query_db("SELECT * FROM transactions WHERE id = ?", [txn_id], one=True)
    if not txn:
        return jsonify({'error': 'Transaksi tidak ditemukan.'}), 404
    items = query_db("SELECT * FROM transaction_items WHERE transaction_id = ?", [txn_id])
    result         = dict(txn)
    result['items']= [dict(i) for i in items]
    return jsonify(result)


@app.route('/api/transactions/<int:txn_id>', methods=['PUT'])
@login_required
def api_transactions_update(txn_id):
    txn = query_db("SELECT * FROM transactions WHERE id = ?", [txn_id], one=True)
    if not txn:
        return jsonify({'error': 'Transaksi tidak ditemukan.'}), 404

    data   = request.json or {}
    action = data.get('action', 'update')

    if action == 'pay':
        if txn['status'] == 'paid':
            return jsonify({'error': 'Transaksi sudah dibayar.'}), 400
        payment_method = data.get('payment_method', 'tunai')
        amount_paid    = float(data.get('amount_paid', txn['total']))
        change_amount  = max(0, amount_paid - txn['total'])
        paid_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        execute_db(
            """UPDATE transactions SET
               status='paid', payment_method=?, amount_paid=?, change_amount=?, paid_at=?
               WHERE id=?""",
            [payment_method, amount_paid, change_amount, paid_at, txn_id]
        )
        # Deduct stock
        items = query_db("SELECT * FROM transaction_items WHERE transaction_id = ?", [txn_id])
        for item in items:
            if item['product_id']:
                prod = query_db("SELECT stock_qty, name FROM products WHERE id = ?", [item['product_id']], one=True)
                if prod:
                    old_qty = prod['stock_qty']
                    new_qty = max(0, old_qty - item['quantity'])
                    execute_db("UPDATE products SET stock_qty = ?, updated_at=CURRENT_TIMESTAMP WHERE id = ?",
                               [new_qty, item['product_id']])
                    execute_db(
                        """INSERT INTO stock_history
                           (product_id, product_name, action, qty_change, qty_before, qty_after, notes, user_id)
                           VALUES (?, ?, 'sale', ?, ?, ?, ?, ?)""",
                        [item['product_id'], prod['name'], -item['quantity'], old_qty, new_qty,
                         f"Terjual via #{txn['order_code']}", session.get('user_id')]
                    )
        return jsonify({'message': 'Pembayaran berhasil.', 'change': change_amount})

    elif action == 'cancel':
        execute_db("UPDATE transactions SET status='cancelled' WHERE id=?", [txn_id])
        return jsonify({'message': 'Transaksi dibatalkan.'})

    elif action == 'update_items':
        if txn['status'] != 'pending':
            return jsonify({'error': 'Hanya transaksi pending yang bisa diubah.'}), 400
        items = data.get('items', [])
        if not items:
            return jsonify({'error': 'Items tidak boleh kosong.'}), 400

        s  = query_db("SELECT key, value FROM settings WHERE key IN ('tax_enabled','tax_rate')")
        sv = {r['key']: r['value'] for r in s}
        tax_enabled = sv.get('tax_enabled', '1') == '1'
        tax_rate    = float(sv.get('tax_rate', '10')) / 100

        subtotal = sum(i.get('price', 0) * i.get('qty', 1) for i in items)
        discount = float(data.get('discount_amount', txn['discount_amount'] or 0))
        taxable  = max(0, subtotal - discount)
        tax_amount = round(taxable * tax_rate, 0) if tax_enabled else 0
        total    = taxable + tax_amount

        execute_db("DELETE FROM transaction_items WHERE transaction_id = ?", [txn_id])
        for item in items:
            execute_db(
                """INSERT INTO transaction_items
                   (transaction_id, product_id, product_name, product_price, quantity, discount_pct, subtotal, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                [txn_id, item.get('product_id'), item.get('name'), item.get('price'),
                 item.get('qty', 1), item.get('discount_pct', 0),
                 item.get('price') * item.get('qty', 1), item.get('notes', '')]
            )
        execute_db(
            "UPDATE transactions SET subtotal=?, discount_amount=?, tax_amount=?, total=?, customer_name=?, table_number=? WHERE id=?",
            [subtotal, discount, tax_amount, total,
             data.get('customer_name', txn['customer_name'] or ''),
             data.get('table_number', txn['table_number'] or ''),
             txn_id]
        )
        return jsonify({'message': 'Pesanan diperbarui.', 'total': total})

    return jsonify({'error': 'Action tidak dikenal.'}), 400


@app.route('/api/transactions/<int:txn_id>', methods=['DELETE'])
@login_required
def api_transaction_delete(txn_id):
    txn = query_db("SELECT * FROM transactions WHERE id = ?", [txn_id], one=True)
    if not txn:
        return jsonify({'error': 'Transaksi tidak ditemukan.'}), 404
    if txn['status'] != 'pending':
        return jsonify({'error': 'Hanya transaksi pending yang dapat dihapus.'}), 400

    execute_db("DELETE FROM transaction_items WHERE transaction_id = ?", [txn_id])
    execute_db("DELETE FROM transactions WHERE id = ?", [txn_id])
    return jsonify({'message': 'Pesanan pending berhasil dihapus.'})


# ─── API: DASHBOARD STATS ───────────────────────────────────────────────────

@app.route('/api/dashboard/stats', methods=['GET'])
@login_required
def api_dashboard_stats():
    period = request.args.get('period', 'today')
    now    = datetime.now()

    if period == 'today':
        from_str  = now.strftime('%Y-%m-%d 00:00:00')
        prev_from = (now - timedelta(days=1)).strftime('%Y-%m-%d 00:00:00')
        prev_to   = from_str
    elif period == 'week':
        from_str  = (now - timedelta(days=7)).strftime('%Y-%m-%d 00:00:00')
        prev_from = (now - timedelta(days=14)).strftime('%Y-%m-%d 00:00:00')
        prev_to   = from_str
    elif period == 'month':
        from_str  = (now - timedelta(days=30)).strftime('%Y-%m-%d 00:00:00')
        prev_from = (now - timedelta(days=60)).strftime('%Y-%m-%d 00:00:00')
        prev_to   = from_str
    else:
        from_str  = '2000-01-01 00:00:00'
        prev_from = '1999-01-01 00:00:00'
        prev_to   = from_str

    curr = query_db(
        "SELECT COALESCE(SUM(total),0) as revenue, COUNT(*) as txn_count FROM transactions WHERE status='paid' AND paid_at >= ?",
        [from_str], one=True
    )
    prev = query_db(
        "SELECT COALESCE(SUM(total),0) as revenue, COUNT(*) as txn_count FROM transactions WHERE status='paid' AND paid_at >= ? AND paid_at < ?",
        [prev_from, prev_to], one=True
    )
    best_seller = query_db(
        """SELECT ti.product_name, SUM(ti.quantity) as total_qty, SUM(ti.subtotal) as total_revenue
           FROM transaction_items ti
           JOIN transactions t ON ti.transaction_id = t.id
           WHERE t.status='paid' AND t.paid_at >= ?
           GROUP BY ti.product_id, ti.product_name
           ORDER BY total_qty DESC LIMIT 1""",
        [from_str], one=True
    )
    low_stock = query_db(
        """SELECT p.id, p.name, p.stock_qty, p.stock_unit, p.low_stock_threshold,
                  CASE WHEN p.stock_qty = 0 THEN 'empty'
                       WHEN p.stock_qty <= p.low_stock_threshold / 2 THEN 'critical'
                       ELSE 'low' END as alert_level
           FROM products p
           WHERE p.stock_qty <= p.low_stock_threshold AND p.is_available = 1
           ORDER BY p.stock_qty ASC LIMIT 15"""
    )

    # Revenue chart (last 7 days)
    chart_data = []
    for i in range(6, -1, -1):
        day = now - timedelta(days=i)
        day_from = day.strftime('%Y-%m-%d 00:00:00')
        day_to   = day.strftime('%Y-%m-%d 23:59:59')
        row = query_db(
            "SELECT COALESCE(SUM(total),0) as rev, COUNT(*) as cnt FROM transactions WHERE status='paid' AND paid_at BETWEEN ? AND ?",
            [day_from, day_to], one=True
        )
        chart_data.append({
            'label': day.strftime('%d/%m'),
            'revenue': row['rev'],
            'count':   row['cnt'],
        })

    curr_rev = curr['revenue'] if curr else 0
    prev_rev = prev['revenue'] if prev else 0
    curr_txn = curr['txn_count'] if curr else 0
    prev_txn = prev['txn_count'] if prev else 0

    rev_growth = round(((curr_rev - prev_rev) / prev_rev) * 100, 1) if prev_rev > 0 else 0
    txn_growth = round(((curr_txn - prev_txn) / prev_txn) * 100, 1) if prev_txn > 0 else 0

    # Average order value
    avg_order = round(curr_rev / curr_txn, 0) if curr_txn > 0 else 0

    return jsonify({
        'revenue':           curr_rev,
        'revenue_growth':    rev_growth,
        'transaction_count': curr_txn,
        'transaction_growth':txn_growth,
        'avg_order_value':   avg_order,
        'best_seller': {
            'name':    best_seller['product_name'] if best_seller else '—',
            'qty':     best_seller['total_qty']    if best_seller else 0,
            'revenue': best_seller['total_revenue']if best_seller else 0,
        },
        'low_stock':   [dict(i) for i in low_stock],
        'chart_data':  chart_data,
    })


# ─── API: LAPORAN ───────────────────────────────────────────────────────────

@app.route('/api/laporan/summary', methods=['GET'])
@admin_required
def api_laporan_summary():
    date_from = request.args.get('date_from', datetime.now().strftime('%Y-%m-%d'))
    date_to   = request.args.get('date_to',   datetime.now().strftime('%Y-%m-%d'))
    from_str  = date_from + ' 00:00:00'
    to_str    = date_to   + ' 23:59:59'

    # Revenue summary
    summary = query_db(
        """SELECT
               COUNT(*) as total_transactions,
               COALESCE(SUM(CASE WHEN status='paid' THEN 1 ELSE 0 END), 0) as paid_count,
               COALESCE(SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END), 0) as cancelled_count,
               COALESCE(SUM(CASE WHEN status='paid' THEN total ELSE 0 END), 0) as gross_revenue,
               COALESCE(SUM(CASE WHEN status='paid' THEN tax_amount ELSE 0 END), 0) as total_tax,
               COALESCE(SUM(CASE WHEN status='paid' THEN discount_amount ELSE 0 END), 0) as total_discount,
               COALESCE(SUM(CASE WHEN status='paid' THEN subtotal ELSE 0 END), 0) as net_revenue
           FROM transactions
           WHERE created_at BETWEEN ? AND ?""",
        [from_str, to_str], one=True
    )

    # Payment method breakdown
    payment_breakdown = query_db(
        """SELECT payment_method, COUNT(*) as count, COALESCE(SUM(total),0) as revenue
           FROM transactions
           WHERE status='paid' AND paid_at BETWEEN ? AND ?
           GROUP BY payment_method""",
        [from_str, to_str]
    )

    # Top products
    top_products = query_db(
        """SELECT ti.product_name, SUM(ti.quantity) as total_qty,
                  SUM(ti.subtotal) as total_revenue,
                  COUNT(DISTINCT ti.transaction_id) as order_count
           FROM transaction_items ti
           JOIN transactions t ON ti.transaction_id = t.id
           WHERE t.status='paid' AND t.paid_at BETWEEN ? AND ?
           GROUP BY ti.product_name
           ORDER BY total_qty DESC LIMIT 20""",
        [from_str, to_str]
    )

    # By cashier
    by_cashier = query_db(
        """SELECT COALESCE(t.cashier_name, u.nama_lengkap, 'Unknown') as kasir,
                  COUNT(*) as txn_count, COALESCE(SUM(t.total),0) as revenue
           FROM transactions t
           LEFT JOIN users u ON t.cashier_id = u.id
           WHERE t.status='paid' AND t.paid_at BETWEEN ? AND ?
           GROUP BY t.cashier_id, kasir
           ORDER BY revenue DESC""",
        [from_str, to_str]
    )

    # Daily breakdown
    daily = query_db(
        """SELECT DATE(paid_at) as tanggal,
                  COUNT(*) as txn_count, COALESCE(SUM(total),0) as revenue
           FROM transactions
           WHERE status='paid' AND paid_at BETWEEN ? AND ?
           GROUP BY DATE(paid_at)
           ORDER BY tanggal""",
        [from_str, to_str]
    )

    return jsonify({
        'period': {'from': date_from, 'to': date_to},
        'summary':           dict(summary) if summary else {},
        'payment_breakdown': [dict(r) for r in payment_breakdown],
        'top_products':      [dict(r) for r in top_products],
        'by_cashier':        [dict(r) for r in by_cashier],
        'daily':             [dict(r) for r in daily],
    })


@app.route('/api/laporan/excel', methods=['GET'])
@admin_required
def api_laporan_excel():
    """Export financial report to Excel."""
    date_from = request.args.get('date_from', datetime.now().strftime('%Y-%m-%d'))
    date_to   = request.args.get('date_to',   datetime.now().strftime('%Y-%m-%d'))
    from_str  = date_from + ' 00:00:00'
    to_str    = date_to   + ' 23:59:59'

    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                      GradientFill, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl belum terinstall. Jalankan: pip install openpyxl'}), 500

    wb = openpyxl.Workbook()

    # ─── Style helpers ───
    def cell_style(ws, row, col, value, bold=False, fill_color=None,
                   font_color='000000', align='left', fmt=None, font_size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, color=font_color, size=font_size)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)
        if fmt:
            cell.number_format = fmt
        thin = Side(style='thin', color='CCCCCC')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return cell

    def set_header_row(ws, row, headers, widths, fill='1B6D24', font_color='FFFFFF'):
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            cell_style(ws, row, col, h, bold=True, fill_color=fill,
                       font_color=font_color, align='center', font_size=10)
            ws.column_dimensions[get_column_letter(col)].width = w

    # Store settings
    store_s = query_db("SELECT key, value FROM settings")
    store   = {r['key']: r['value'] for r in store_s}
    nama_usaha = store.get('nama_usaha', 'Warung Penyetan Hoki')

    # ─── Sheet 1: Ringkasan ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Ringkasan'
    ws1.row_dimensions[1].height = 35

    # Title
    ws1.merge_cells('A1:F1')
    title_cell = ws1['A1']
    title_cell.value = f'LAPORAN KEUANGAN — {nama_usaha}'
    title_cell.font  = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill  = PatternFill('solid', fgColor='1B6D24')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('A2:F2')
    period_cell = ws1['A2']
    period_cell.value = f'Periode: {date_from} s/d {date_to}'
    period_cell.font  = Font(size=11, color='666666', italic=True)
    period_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 22

    summary = query_db(
        """SELECT
               COALESCE(SUM(CASE WHEN status='paid' THEN 1 ELSE 0 END), 0) as paid_count,
               COALESCE(SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END), 0) as cancelled_count,
               COALESCE(SUM(CASE WHEN status='paid' THEN total ELSE 0 END), 0) as gross_revenue,
               COALESCE(SUM(CASE WHEN status='paid' THEN tax_amount ELSE 0 END), 0) as total_tax,
               COALESCE(SUM(CASE WHEN status='paid' THEN discount_amount ELSE 0 END), 0) as total_discount,
               COALESCE(SUM(CASE WHEN status='paid' THEN subtotal ELSE 0 END), 0) as net_revenue,
               COALESCE(AVG(CASE WHEN status='paid' THEN total ELSE NULL END), 0) as avg_order
           FROM transactions
           WHERE created_at BETWEEN ? AND ?""",
        [from_str, to_str], one=True
    )
    s = dict(summary) if summary else {}

    summary_rows = [
        ('', '', '', '', '', ''),
        ('RINGKASAN PENDAPATAN', '', '', '', '', ''),
        ('Metrik', 'Nilai', '', '', '', ''),
        ('Total Transaksi Selesai', s.get('paid_count', 0)),
        ('Total Transaksi Batal',   s.get('cancelled_count', 0)),
        ('Gross Revenue',           s.get('gross_revenue', 0)),
        ('Total Diskon',            s.get('total_discount', 0)),
        ('Net Revenue (setelah diskon)', max(0, s.get('gross_revenue', 0) - s.get('total_discount', 0))),
        ('Total Pajak (PPN)',        s.get('total_tax', 0)),
        ('Rata-rata Nilai Transaksi',s.get('avg_order', 0)),
    ]

    row = 4
    cell_style(ws1, row, 1, 'METRIK', bold=True, fill_color='2E7D32', font_color='FFFFFF', align='center')
    cell_style(ws1, row, 2, 'NILAI', bold=True, fill_color='2E7D32', font_color='FFFFFF', align='center')
    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 22
    row += 1

    money_fmt = '#,##0'
    for i, (label, value) in enumerate(summary_rows[3:], start=row):
        fill = 'F4FAF4' if i % 2 == 0 else 'FFFFFF'
        cell_style(ws1, i, 1, label, fill_color=fill)
        is_money = isinstance(value, float) or (isinstance(value, int) and i > row + 1)
        cell_style(ws1, i, 2, value, fill_color=fill, align='right',
                   fmt=money_fmt if is_money else None)

    # Payment breakdown
    pay_row = row + len(summary_rows) - 3 + 2
    cell_style(ws1, pay_row, 1, 'METODE PEMBAYARAN', bold=True, fill_color='005FAF',
               font_color='FFFFFF', align='center')
    cell_style(ws1, pay_row, 2, 'JUMLAH TXN', bold=True, fill_color='005FAF',
               font_color='FFFFFF', align='center')
    cell_style(ws1, pay_row, 3, 'TOTAL REVENUE', bold=True, fill_color='005FAF',
               font_color='FFFFFF', align='center')
    ws1.column_dimensions['C'].width = 22
    pay_row += 1
    pay_data = query_db(
        """SELECT COALESCE(payment_method,'—') as payment_method, COUNT(*) as count,
                  COALESCE(SUM(total),0) as revenue
           FROM transactions WHERE status='paid' AND paid_at BETWEEN ? AND ?
           GROUP BY payment_method""",
        [from_str, to_str]
    )
    for i, p in enumerate(pay_data):
        fill = 'EEF4FF' if i % 2 == 0 else 'FFFFFF'
        cell_style(ws1, pay_row, 1, p['payment_method'].upper(), fill_color=fill)
        cell_style(ws1, pay_row, 2, p['count'], fill_color=fill, align='center')
        cell_style(ws1, pay_row, 3, p['revenue'], fill_color=fill, align='right', fmt=money_fmt)
        pay_row += 1

    # ─── Sheet 2: Detail Transaksi ───────────────────────────────────────
    ws2 = wb.create_sheet('Detail Transaksi')
    ws2.merge_cells('A1:I1')
    ws2['A1'].value = f'DETAIL TRANSAKSI — {date_from} s/d {date_to}'
    ws2['A1'].font  = Font(bold=True, size=13, color='FFFFFF')
    ws2['A1'].fill  = PatternFill('solid', fgColor='1B6D24')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30

    headers = ['No.','Kode Order','Waktu','Kasir','Pelanggan','Meja','Subtotal','Diskon','Pajak','Total','Metode','Status']
    widths  = [5, 14, 18, 18, 18, 10, 16, 14, 14, 16, 14, 12]
    set_header_row(ws2, 2, headers, widths)

    txns = query_db(
        """SELECT t.*, COALESCE(t.cashier_name, u.nama_lengkap, '—') as kasir_name
           FROM transactions t LEFT JOIN users u ON t.cashier_id = u.id
           WHERE t.created_at BETWEEN ? AND ?
           ORDER BY t.created_at""",
        [from_str, to_str]
    )

    for i, t in enumerate(txns, start=1):
        row_n = i + 2
        fill  = 'F8FFF8' if i % 2 == 0 else 'FFFFFF'
        status_fill = {'paid': 'E8F5E9', 'cancelled': 'FFEBEE', 'pending': 'FFF8E1'}.get(t['status'], 'FFFFFF')
        cols = [
            (i,                      None),
            (t['order_code'],        None),
            (t['created_at'][:16],   None),
            (t['kasir_name'],        None),
            (t['customer_name'] or '—', None),
            (t['table_number'] or '—',  None),
            (t['subtotal'],          money_fmt),
            (t['discount_amount'] or 0, money_fmt),
            (t['tax_amount'],        money_fmt),
            (t['total'],             money_fmt),
            ((t['payment_method'] or '—').upper(), None),
            (t['status'].upper(),    None),
        ]
        for col, (val, fmt) in enumerate(cols, start=1):
            f = status_fill if col == 12 else fill
            cell_style(ws2, row_n, col, val, fill_color=f, fmt=fmt,
                       align='right' if fmt else ('center' if col in [1,11,12] else 'left'))

    # ─── Sheet 3: Produk Terlaris ────────────────────────────────────────
    ws3 = wb.create_sheet('Produk Terlaris')
    ws3.merge_cells('A1:F1')
    ws3['A1'].value = f'PRODUK TERLARIS — {date_from} s/d {date_to}'
    ws3['A1'].font  = Font(bold=True, size=13, color='FFFFFF')
    ws3['A1'].fill  = PatternFill('solid', fgColor='1B6D24')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 30

    prod_headers = ['No.', 'Nama Produk', 'Total Terjual', 'Jumlah Order', 'Total Revenue', 'Rata-rata/Order']
    prod_widths  = [5, 30, 16, 16, 18, 18]
    set_header_row(ws3, 2, prod_headers, prod_widths)

    top_prods = query_db(
        """SELECT ti.product_name, SUM(ti.quantity) as total_qty,
                  COUNT(DISTINCT ti.transaction_id) as order_count,
                  SUM(ti.subtotal) as total_revenue
           FROM transaction_items ti
           JOIN transactions t ON ti.transaction_id = t.id
           WHERE t.status='paid' AND t.paid_at BETWEEN ? AND ?
           GROUP BY ti.product_name
           ORDER BY total_qty DESC""",
        [from_str, to_str]
    )
    for i, p in enumerate(top_prods, start=1):
        row_n = i + 2
        fill  = 'FFF9E6' if i == 1 else ('F9FFF9' if i % 2 == 0 else 'FFFFFF')
        avg = round(p['total_revenue'] / p['order_count'], 0) if p['order_count'] else 0
        row_data = [
            (i,                   None,       'center'),
            (p['product_name'],   None,       'left'),
            (p['total_qty'],      '#,##0',    'center'),
            (p['order_count'],    '#,##0',    'center'),
            (p['total_revenue'],  money_fmt,  'right'),
            (avg,                 money_fmt,  'right'),
        ]
        for col, (val, fmt, align) in enumerate(row_data, start=1):
            bold = i == 1 and col in [2, 3, 5]
            cell_style(ws3, row_n, col, val, fill_color=fill, fmt=fmt, align=align, bold=bold)

    # ─── Sheet 4: Riwayat Stok ──────────────────────────────────────────
    ws4 = wb.create_sheet('Riwayat Stok')
    ws4.merge_cells('A1:G1')
    ws4['A1'].value = f'RIWAYAT STOK — {date_from} s/d {date_to}'
    ws4['A1'].font  = Font(bold=True, size=13, color='FFFFFF')
    ws4['A1'].fill  = PatternFill('solid', fgColor='1B6D24')
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 30

    stock_headers = ['No.','Waktu','Produk','Aksi','Sebelum','Perubahan','Sesudah','Keterangan']
    stock_widths  = [5, 18, 28, 14, 10, 12, 10, 30]
    set_header_row(ws4, 2, stock_headers, stock_widths)

    stock_hist = query_db(
        """SELECT sh.*, u.nama_lengkap as user_name
           FROM stock_history sh LEFT JOIN users u ON sh.user_id = u.id
           WHERE DATE(sh.created_at) BETWEEN ? AND ?
           ORDER BY sh.created_at DESC LIMIT 500""",
        [date_from, date_to]
    )
    action_colors = {'restock': 'E8F5E9', 'sale': 'FFEBEE', 'adjustment': 'FFF8E1', 'initial': 'EEF4FF'}
    for i, sh in enumerate(stock_hist, start=1):
        row_n = i + 2
        fill  = action_colors.get(sh['action'], 'FFFFFF')
        cols  = [
            (i,                sh['action'],  sh['product_name'],
             sh['action'].upper(), sh['qty_before'],
             sh['qty_change'], sh['qty_after'], sh['notes'] or '—')
        ]
        row_data = [i, sh['created_at'][:16], sh['product_name'],
                    sh['action'].upper(), sh['qty_before'],
                    ('+' if sh['qty_change'] >= 0 else '') + str(sh['qty_change']),
                    sh['qty_after'], sh['notes'] or '—']
        for col, val in enumerate(row_data, start=1):
            cell_style(ws4, row_n, col, val, fill_color=fill,
                       align='right' if col in [5, 7] else ('center' if col in [1, 4, 6] else 'left'))

    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"laporan_{date_from}_sd_{date_to}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ─── API: SETTINGS ──────────────────────────────────────────────────────────

@app.route('/api/settings', methods=['GET'])
@login_required
def api_settings_get():
    rows = query_db("SELECT key, value FROM settings")
    result = {}
    for row in rows:
        # Don't send media blobs in general settings call (too large)
        if row['key'] in ('logo_data', 'banner_data', 'qris_data'):
            result[row['key']] = '1' if row['value'] else '0'  # Just indicate presence
        else:
            result[row['key']] = row['value']
    return jsonify(result)


@app.route('/api/settings', methods=['PUT'])
@admin_required
def api_settings_update():
    data = request.json or {}
    for key, value in data.items():
        # Skip media keys via JSON (they use separate endpoints)
        if key in ('logo_data', 'banner_data', 'qris_data'):
            continue
        execute_db(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            [key, str(value)]
        )
    return jsonify({'message': 'Pengaturan berhasil disimpan.'})


@app.route('/api/settings/media/<media_type>', methods=['GET'])
@login_required
def api_settings_media_get(media_type):
    """Get media (logo, banner, qris) as data URL."""
    if media_type not in ('logo_data', 'banner_data', 'qris_data'):
        return jsonify({'error': 'Tipe media tidak valid.'}), 400
    row = query_db("SELECT value FROM settings WHERE key = ?", [media_type], one=True)
    if not row or not row['value']:
        return jsonify({'data': None})
    return jsonify({'data': row['value']})


@app.route('/api/settings/media/<media_type>', methods=['POST'])
@admin_required
def api_settings_media_upload(media_type):
    """Upload media (logo, banner, qris) and store as base64."""
    if media_type not in ('logo_data', 'banner_data', 'qris_data'):
        return jsonify({'error': 'Tipe media tidak valid.'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'File tidak ditemukan.'}), 400
    file = request.files['file']
    if not file or not file.filename or not allowed_image(file.filename):
        return jsonify({'error': 'Format file tidak didukung.'}), 400

    # Different max sizes for different media
    max_sizes = {
        'logo_data':   (400, 400),
        'banner_data': (1200, 400),
        'qris_data':   (600, 600),
    }
    data_url = encode_image(file, max_size=max_sizes[media_type])
    execute_db(
        "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
        [media_type, data_url]
    )
    return jsonify({'message': 'Media berhasil diupload.', 'data': data_url})


@app.route('/api/settings/media/<media_type>', methods=['DELETE'])
@admin_required
def api_settings_media_delete(media_type):
    if media_type not in ('logo_data', 'banner_data', 'qris_data'):
        return jsonify({'error': 'Tipe media tidak valid.'}), 400
    execute_db("UPDATE settings SET value='' WHERE key=?", [media_type])
    return jsonify({'message': 'Media dihapus.'})


# ─── API: USERS ─────────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
@admin_required
def api_users():
    users = query_db(
        "SELECT id, username, role, nama_lengkap, is_active, avatar_color, created_at FROM users ORDER BY id"
    )
    return jsonify([dict(u) for u in users])


@app.route('/api/users', methods=['POST'])
@admin_required
def api_users_create():
    data      = request.json or {}
    username  = data.get('username', '').strip()
    password  = data.get('password', '').strip()
    role      = data.get('role', 'kasir')
    nama      = data.get('nama_lengkap', '').strip()
    color     = data.get('avatar_color', '#0d631b')
    if not username or not password:
        return jsonify({'error': 'Username dan password wajib diisi.'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Password minimal 6 karakter.'}), 400
    existing = query_db("SELECT id FROM users WHERE username = ?", [username], one=True)
    if existing:
        return jsonify({'error': 'Username sudah digunakan.'}), 409
    user_id = execute_db(
        "INSERT INTO users (username, password_hash, role, nama_lengkap, avatar_color) VALUES (?, ?, ?, ?, ?)",
        [username, generate_password_hash(password), role, nama, color]
    )
    return jsonify({'id': user_id, 'username': username, 'role': role}), 201


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def api_users_update(user_id):
    data     = request.json or {}
    existing = query_db("SELECT * FROM users WHERE id = ?", [user_id], one=True)
    if not existing:
        return jsonify({'error': 'User tidak ditemukan.'}), 404
    nama      = data.get('nama_lengkap', existing['nama_lengkap'])
    role      = data.get('role',         existing['role'])
    is_active = int(data.get('is_active', existing['is_active']))
    color     = data.get('avatar_color', existing['avatar_color'] or '#0d631b')
    if data.get('password'):
        if len(data['password']) < 6:
            return jsonify({'error': 'Password minimal 6 karakter.'}), 400
        execute_db(
            "UPDATE users SET nama_lengkap=?, role=?, is_active=?, password_hash=?, avatar_color=? WHERE id=?",
            [nama, role, is_active, generate_password_hash(data['password']), color, user_id]
        )
    else:
        execute_db(
            "UPDATE users SET nama_lengkap=?, role=?, is_active=?, avatar_color=? WHERE id=?",
            [nama, role, is_active, color, user_id]
        )
    return jsonify({'message': 'User berhasil diperbarui.'})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_users_delete(user_id):
    if user_id == session.get('user_id'):
        return jsonify({'error': 'Tidak dapat menghapus akun sendiri.'}), 400
    execute_db("DELETE FROM users WHERE id = ?", [user_id])
    return jsonify({'message': 'User berhasil dihapus.'})


# ─── API: STOCK HISTORY ──────────────────────────────────────────────────────

@app.route('/api/stock/history', methods=['GET'])
@login_required
def api_stock_history():
    product_id = request.args.get('product_id')
    limit      = int(request.args.get('limit', 50))
    query = """
        SELECT sh.*, u.nama_lengkap as user_name
        FROM stock_history sh LEFT JOIN users u ON sh.user_id = u.id
        WHERE 1=1
    """
    args = []
    if product_id:
        query += " AND sh.product_id = ?"
        args.append(product_id)
    query += " ORDER BY sh.created_at DESC LIMIT ?"
    args.append(limit)
    rows = query_db(query, args)
    return jsonify([dict(r) for r in rows])


# ─── API: BACKUP ────────────────────────────────────────────────────────────

@app.route('/api/backup', methods=['GET'])
@admin_required
def api_backup():
    from database import DATABASE
    if not os.path.exists(DATABASE):
        return jsonify({'error': 'Database belum ada.'}), 404
    timestamp     = datetime.now().strftime('%Y%m%d_%H%M%S')
    download_name = f'hokiapp_backup_{timestamp}.db'
    return send_file(DATABASE, as_attachment=True, download_name=download_name,
                     mimetype='application/octet-stream')


# ─── MAIN ───────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    print("=" * 55)
    print("  [POS] HokiApp POS v2.0 - Warung Penyetan Hoki")
    print("  Server: http://localhost:5000")
    print("  Login Admin:  admin   / hoki2024")
    print("  Login Kasir:  kasir   / kasir123")
    print("=" * 55)
    app.run(debug=True, host='localhost', port=5000)
